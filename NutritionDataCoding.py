# coding: utf-8

# In[1]:

import openpyxl, re
from openpyxl import Workbook


# #### Here is a various assortment of helpful methods that will be called by the main loop. We modularize in an attempt to clean up the code and make it more readable.

# In[ ]:

def fill_nutrition_data():
    # get all the numerical nutrition data into the new sheet
    nutrition_data = [sheet.cell(row=i+1, column=j).value for j in range(2,9)]
    
    for j in range(6,13):
        structured_sheet.cell(row=structured_sheet_pointer+1, column=j).value = nutrition_data[j-6]
        
    # clear out this row in the structured sheet for new data coming in
        for j in range(1, 6):
            structured_sheet.cell(row=structured_sheet_pointer+1, column=j).value = ''
            
def get_product_category():
    # find and store the product category (each entry must have one)
    for j in range(i+1, 0, -1): # step backwards through the rows of the spreadsheet
        # grab data in current cell, in the first column
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        # if the cell is bold, we know that it contains product category information
        if current_cell.style.font.bold and '(cont.)' not in current_cell.value and current_cell.value[-2:] != ' .' and current_cell.value[0] != ' ':
            # set the first column in the current row of the restructured sheet to this product category
            product_category = re.sub('\(.*\)', '', current_cell.value.split(',')[0])
            product_category = product_category.strip().replace(':', '').replace('.', '').replace('_', '').replace('"', '')
            product_category = product_category.replace(u'\u201c', '').replace(u'\u201d', '')
            product_category = product_category.split('(')[0]
            structured_sheet.cell(row=structured_sheet_pointer+1, column=1).value = product_category
            if j not in product_category_rows:
                product_category_rows.append(j) # add this 
            break

def get_product_description():
    # find and store the product description (if it exists)
    for j in range(i+1, 0, -1): # step backwards through the rows of the spreadsheet
        # grab data in current cell, in the first column
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        if current_cell.style.font.bold:
            cell_text = re.sub('\(.*\)', '', current_cell.value) # remove all parenthesized text
            cell_tokens = cell_text.split(',') # split the cell's text into chunks, separated by comma occurence
            product_description = get_description(cell_tokens[1:]) # outsource implementation to method (remove chunk before 1st comma)
            if not product_description: # if there is no text left over
                break
            product_description = product_description.replace('.', '').replace('_', '')
            structured_sheet.cell(row=structured_sheet_pointer+1, column=2).value = product_description
            if j not in product_description_rows:
                product_description_rows.append(j)
            break
        # if the cell contains some text preceding a semicolon, it must contain a product description
        if ':' not in current_cell.value:
            continue
        else: # there must be a product description
            cell_text = re.sub('\(.*\)', '', current_cell.value) # remove all parenthesized text
            cell_tokens = cell_text.split(',') # split the cell's text into chunks, separated by comma occurence
            product_description = get_description(cell_tokens) # outsource implementation to method
            if not product_description: # if there is no text left over
                continue
            product_description = product_description.replace('.', '').replace('_', '')
            structured_sheet.cell(row=structured_sheet_pointer+1, column=2).value = product_description
            if j not in product_description_rows:
                product_description_rows.append(j)
            break

    if not structured_sheet.cell(row=structured_sheet_pointer+1, column=2).value:
        structured_sheet.cell(row=structured_sheet_pointer+1, column=2).value = 'NA'

def get_description(tokens):
    ret = ''
    for token in tokens:
        if any(char.isdigit() for char in token) or len(token) == 0 or 'except as noted' in token:
            continue
        else:
            if ret == '':
                ret = token.strip()
            else:
                ret += ', ' + token.strip()
    return ret.strip().strip(':')

def get_brand_name():
    # find and store the brand name (if it exists)
    for j in range(i+1, 0, -1):
        # grab data in current cell, in the first column            
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        if j+1 in brand_name_rows:
            break
        # if the cell contains parenthesized text (not equal to cont.), it must contain brand name information
        if current_cell.value.find('(') != -1 and '(cont.)' not in current_cell.value and 'see also' not in current_cell.value:
            # get the cell's text within parentheses; this is the brand name
            brand_name = current_cell.value[current_cell.value.find("(")+1:current_cell.value.find(")")]
            brand_name = brand_name.replace('.', '').replace('_', '')
            structured_sheet.cell(row=structured_sheet_pointer+1, column=4).value = brand_name
            if j not in brand_name_rows:
                brand_name_rows.append(j)
            break

        elif j in product_category_rows or j in product_description_rows:
            # there cannot be any brand name information before here
            structured_sheet.cell(row=structured_sheet_pointer+1, column=4).value = 'NA'
            break

    if not structured_sheet.cell(row=structured_sheet_pointer+1, column=4).value.strip():
        structured_sheet.cell(row=structured_sheet_pointer+1, column=4).value = 'NA'

def get_type():
    # find and store the type (if it exists)
    for j in range(i+1, 0, -1): # check current and all previous rows up to the last product category row
        # grab data in current cell, in the first column
        current_cell = sheet.cell(row=j, column=1)            
        if not current_cell.value:
            continue
        if j in product_category_rows or j+1 in brand_name_rows:
            break
        # if the cell contains parentheses
        
        if '(' in current_cell.value and j in brand_name_rows and '(cont.)' not in current_cell.value:
            # get the token with parentheses
            token = current_cell.value.split('(')[0]
            if token != '' and ':' not in token:
                token = token.strip().replace('.', '').replace('_', '')
                type_ = remove_numerics(token)
                structured_sheet.cell(row=structured_sheet_pointer+1, column=3).value = type_
                break
        # if the cell doesn't contain parentheses and doesn't contain a semicolon
        elif ':' not in current_cell.value:
            type_ = current_cell.value.replace('.', '').replace('_', '').strip()
            type_ = remove_numerics(type_)
            structured_sheet.cell(row=structured_sheet_pointer+1, column=3).value = type_
            break

    if not structured_sheet.cell(row=structured_sheet_pointer+1, column=3).value.strip():
        structured_sheet.cell(row=structured_sheet_pointer+1, column=3).value = 'NA'

def remove_numerics(token):
    ret = ''
    for x in token.split(','):
        if not any(char.isdigit() for char in x):
            ret += x + ','
    ret = ret.strip().strip(',')
    return ret

def get_serving_size():
    # find and store the serving size (each entry must have one)
    for j in range(i+1, 0, -1): # check all previous rows up to the last product category row
        # grab data in current cell, in the first column            
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        # if the cell contains some numerical data, it will pertain to the serving size
        if any(char.isdigit() for char in current_cell.value):
            # split the cell's text by comma, looking for the token with numerical data
            serving_size = format_serving_size(current_cell.value.split(','))

            if serving_size:
                serving_size = serving_size.replace('.', '').replace('_', '').replace('*', '').strip()
                structured_sheet.cell(row=structured_sheet_pointer+1, column=5).value = serving_size
                break
            else:
                continue

        elif j in product_description_rows or j in product_category_rows:
            # split the cell's text by comma, looking for the token with numerical data
            serving_size = format_serving_size(current_cell.value.split(','))

            if serving_size:
                serving_size = serving_size.replace('.', '').replace('_', '').replace('*', '').strip()
                structured_sheet.cell(row=structured_sheet_pointer+1, column=5).value = serving_size
                break
            else:
                continue

def format_serving_size(tokens):
    for token in tokens: # loop through each token
        if any(char.isdigit() for char in token) and '(' not in token and u'\N{DEGREE SIGN}' not in token: # looking for the token with the serving size data...
            token = token.replace('.', '').strip().strip(':')
            return token.split(':')[0]
        elif any(char.isdigit() for char in token.split(':')[0]) and '(' not in token.split(':')[0] and u'\N{DEGREE SIGN}' not in token:
            return token.split(':')[0].strip()


# #### Here we have the main loop. For each row which contains nutrition information, we in turn must find the associated (1) Product Category, (2) Serving Size, and try to find its (3) Product Description, (4) Type, and (5) Brand Name, if they exist for the row's entry.

# In[ ]:

# file_name = input('Enter in the filename, in single or double quotes, of the workbook you\'d like to modify: ')

file_name = raw_input('Enter filename of workbook to be modified: ')
book = openpyxl.load_workbook(file_name)
worksheet = raw_input('Enter name of worksheet to work on during this session: ')
sheet = book[worksheet]

import os.path

if not os.path.isfile('restructured_' + file_name):
    structured_book = Workbook()
    structured_book.remove_sheet(structured_book.get_sheet_by_name('Sheet'))
    for sheet_name in book.get_sheet_names():
        structured_book.create_sheet(sheet_name)
else:
    structured_book = openpyxl.load_workbook('restructured_' + file_name)

for cur_sheet in book:
    if cur_sheet.title == worksheet:
        to_delete = structured_book.get_sheet_by_name(cur_sheet.title)
        structured_book.remove_sheet(to_delete)
        break

structured_sheet = structured_book.create_sheet(worksheet)

# pointers to rows which contain these types of data points
product_category_rows = []
product_description_rows = []
brand_name_rows = []

# a pointer to where we are in the newly formatted spreadsheet
structured_sheet_pointer = 0

# the main loop
for i, row in enumerate(sheet): # check each row for nutrition data
    if i >= len(sheet.rows): # in case there are a lot of blank rows at the end of a sheet (this is an anomaly in the first workbook)
        break
    if i % 500 == 0:
        print 'Progress: ' + str(i) + ' / ' + str(len(sheet.rows))
    if i == len(sheet.rows) - 1:
        print 'Progress: ' + str(i+1) + ' / ' + str(len(sheet.rows))
    if row[1].value: # nutrition data always starts in row 2. if there is data in row 2, there is data in rows 2-8

        fill_nutrition_data()

        get_product_category()
        get_product_description()
        get_brand_name()
        get_type()
        get_serving_size()

        structured_sheet_pointer += 1
    
structured_book.save('restructured_' + file_name)


# In[ ]:

# sets widths of spreadsheet columns based on longest column entry
column_widths = []
for j, row in enumerate(structured_sheet):
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            try:
                if len(cell.value) > column_widths[i]:
                    column_widths[i] = len(cell.value)
            except TypeError:
                pass
        else:
            try:
                column_widths += [len(cell.value)]
            except TypeError:
                pass

structured_sheet.column_dimensions["A"].width = column_widths[0] + 1.0
structured_sheet.column_dimensions["B"].width = column_widths[1] + 1.0
structured_sheet.column_dimensions["C"].width = column_widths[2] + 1.0
structured_sheet.column_dimensions["D"].width = column_widths[3] + 1.0
structured_sheet.column_dimensions["E"].width = column_widths[4] + 1.0
structured_sheet.column_dimensions["F"].width = 6.0
structured_sheet.column_dimensions["G"].width = 6.0
structured_sheet.column_dimensions["H"].width = 6.0
structured_sheet.column_dimensions["I"].width = 6.0
structured_sheet.column_dimensions["J"].width = 6.0
structured_sheet.column_dimensions["K"].width = 6.0
structured_sheet.column_dimensions["L"].width = 6.0

print 'Saving restructured worksheet in workbook: restructured_' + file_name

structured_book.save('restructured_' + file_name)


# In[ ]:



