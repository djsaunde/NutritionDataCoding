'''
This script is meant to serve to set up the workbook / worksheet the user wishes to work with, and then executes the main loop,
restructuring the data into the new workbook, all the while calling helper methods from 'util.py'.

author: Dan Saunders (djsaunde@umass.edu)
'''

# importing packages for spreadsheet manipulation, string matching, file system functions, and warning functions
import openpyxl, re, os.path, warnings
# import all helper methods from util.py
from util import *

# ignore all warnings (super annoying being printed to the user!)
warnings.filterwarnings("ignore")


print '\n'

# getting the name of the file to modify
file_name = ''
while not file_name:
    # prompt the user for a file name
    file_name = raw_input('Input the name of the workbook ("Enter" for previous workbook): ')
    # if the user hasn't entered a file name...
    if not file_name:
        # check for stored file name
        with open('../documents/previous_book.txt', 'r') as f:
            file_name = f.readlines()[0]
            # if no filename is stored, try again from the top of the loop
            if not file_name:
                continue

# opening the workbook using the openpyxl package                
book = openpyxl.load_workbook('../data/' + file_name)

# rewriting previous file name file
with open('../documents/previous_book.txt', 'w') as f:
    f.truncate()
    f.write(file_name)

    
print '\n'

# getting the name of the worksheet to modify
worksheet = ''
while not worksheet:
    # prompt the user for a worksheet title
    worksheet = raw_input('Input the title of the worksheet ("Enter" for previous worksheet): ')
    # if the user hasn't entered a worksheet title...
    if not worksheet:
        # check for stored worksheet title
        with open('../documents/previous_sheet.txt', 'r') as f:
            worksheet = f.readlines()[0]
            # if no worksheet title is stored, try again from the top of the loop
            if not worksheet:
                continue
                
# indexing the sheet from the open workbook into the variable 'sheet'                
sheet = book[worksheet]

# rewriting previous sheet title file
with open('../documents/previous_sheet.txt', 'w') as f:
    f.truncate()
    f.write(worksheet)
    

print '\n'

# if the file for the restructured output doesn't yet exist...
if not os.path.isfile('../data/restructured_' + file_name):
    # we create a new workbook object
    structured_book = openpyxl.Workbook()
    # we remove the default worksheet the workbook is created with
    structured_book.remove_sheet(structured_book.get_sheet_by_name('Sheet'))
    # we create new sheets for each of the sheets in the workbook we are restructuring
    for sheet_name in book.get_sheet_names():
        structured_book.create_sheet(sheet_name)
# otherwise, the file for the restructured output already exists
else:
    # so we load up the workbook which is meant to contain the restructured data
    structured_book = openpyxl.load_workbook('../data/restructured_' + file_name)


# for each sheet in the workbook to be restructured...
for cur_sheet in book:
    # we look for the sheet we are working on in the current session
    if cur_sheet.title == worksheet:
        # we get the sheet by its title
        to_delete = structured_book.get_sheet_by_name(cur_sheet.title)
        # and delete it!
        structured_book.remove_sheet(to_delete)
        break

# finally, we recreate the sheet with the title the user entered
structured_sheet = structured_book.create_sheet(worksheet)

# a pointer to where we are in the newly formatted spreadsheet
cur_pos = 0

# we go through each row and apply our helper methods when applicable to reformat the data into the new spreadsheet
# check each row for nutrition data
for i, row in enumerate(sheet): 
    # in case there are a lot of blank rows at the end of a sheet
    if i >= len(sheet.rows): 
        break
    # keeping track of progress through the restructuring
    if i % 500 == 0:
        print 'Progress: ' + str(i) + ' / ' + str(len(sheet.rows))
    # if we've reached the very last row, let the user know
    if i == len(sheet.rows) - 1:
        print 'Progress: ' + str(i+1) + ' / ' + str(len(sheet.rows))
    # nutrition data always starts in row 2. if there is data in row 2, there is data in rows 2-8
    if row[1].value: 
        # call helper methods in order
        # this grabs the nutrition data only
        fill_nutrition_data(sheet, structured_sheet, cur_pos, i)
        # these grab, in turn, the five pieces of information laid out in the project specification
        structured_sheet = get_product_category(sheet, structured_sheet, cur_pos, i)
        structured_sheet = get_product_description(sheet, structured_sheet, cur_pos, i)
        structured_sheet = get_brand_name(sheet, structured_sheet, cur_pos, i)
        structured_sheet = get_type(sheet, structured_sheet, cur_pos, i)
        structured_sheet = get_serving_size(sheet, structured_sheet, cur_pos, i)
        # increment the spreadsheet pointer to keep track of where we are
        cur_pos += 1
    
# save the workbook after we've completed restructuring
structured_book.save('../data/restructured_' + file_name)


# sets widths of spreadsheet columns based on longest column entry
column_widths = []
for j, row in enumerate(structured_sheet):
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            try:
                if len(cell.value) > column_widths[i]:
                    column_widths[i] = len(cell.value)
            except TypeError:
                pass
        else:
            try:
                column_widths += [len(cell.value)]
            except TypeError:
                pass

for i, char in enumerate(['A', 'B', 'C', 'D', 'E']):
    structured_sheet.column_dimensions[char].width = column_widths[i] + 1.0

for char in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
    structured_sheet.column_dimensions[char].width = 6.0


# save the workbook after we've set column widths according to their contents
structured_book.save('../data/restructured_' + file_name)

