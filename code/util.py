'''
This script is meant to contain helper methods for the main script, make_sheet.py, in an effort to modularize / clean up the
development process. This module will be imported into the main script so as to use the helper methods as needed.
'''

import openpyxl, re


# pointers to rows which contain these types of data points (Product Category, Product Description, or Brand Name)
pc_rows = []
pd_rows = []
bn_rows = []


def fill_nutrition_data(sheet, structured_sheet, cur_pos, i):
    # get all the numerical nutrition data into the new sheet
    nutrition_data = [sheet.cell(row=i+1, column=j).value for j in range(2,9)]
    
    for j in range(6,13):
        structured_sheet.cell(row=cur_pos+1, column=j).value = nutrition_data[j-6]
        
        # clear out this row in the structured sheet for new data coming in
        for j in range(1, 6):
            structured_sheet.cell(row=cur_pos+1, column=j).value = ''

            
def get_product_category(sheet, structured_sheet, cur_pos, i):
    # find and store the product category (each entry must have one)
    for j in range(i+1, 0, -1): # step backwards through the rows of the spreadsheet
        # grab data in current cell, in the first column
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        # if the cell is bold, we know that it contains product category information
        if current_cell.style.font.bold and '(cont.)' not in current_cell.value and current_cell.value[-2:] != ' .' and current_cell.value[0] != ' ':
            product_category = ''
            if 'see' in current_cell.value:
                for token in current_cell.value.split(u'\u201d'):
                    if 'see' in token:
                        if '(' in token:
                            product_category += token.split('(')[0]
                            break
                        continue
                    else:
                        product_category += token
                        break
            else:
                product_category = current_cell.value
            if not product_category: # no text was returned by above logic
                continue             # move on to the previous row
            product_category = re.sub('\(.*\)', '', product_category.split(',')[0])
            if ':' in product_category:
                product_category = product_category.split(':')[0]
            product_category = product_category.strip().replace(')', '').replace(':', '').replace('.', '').replace('_', '').replace('"', '')
            product_category = product_category.replace(u'\u201c', '').replace(u'\u201d', '').replace(u'\u002A', '')
            product_category = product_category.split('(')[0]
            structured_sheet.cell(row=cur_pos+1, column=1).value = product_category
            if j not in pc_rows:
                pc_rows.append(j) # add this 
            break

            
def get_product_description(sheet, structured_sheet, cur_pos, i):
    # find and store the product description (if it exists)
    for j in range(i+1, 0, -1): # step backwards through the rows of the spreadsheet
        # grab data in current cell, in the first column
        current_cell = sheet.cell(row=j, column=1)
        # if there isn't any text in the cell, we skip it
        if not current_cell.value:
            continue
        # if the cell contains bold text, its a product category row
        # we check the cell for product description text as well
        if current_cell.style.font.bold:
            cell_text = ''
            # logic to handle deliquent tokens containing "see"
            if 'see' in current_cell.value:
                # split cell text into tokens separted by right quotation marks (see "..." ...)
                for token in current_cell.value.split(u'\u201d'):
                    # ignore tokens with "see" in them
                    if 'see' in token:
                        continue
                    # add the current token to the total valid cell text
                    else:
                        cell_text += token
            # if there are no occurrences of "see", take all the text in the cell
            else:
                cell_text = current_cell.value
                
            # remove all parenthesized text
            cell_text = re.sub('\(.*\)', '', cell_text)
            # split the cell's text into chunks, separated by comma occurence
            cell_tokens = cell_text.split(',')
            # outsource implementation to method (remove chunk before 1st comma)
            product_description = get_description(cell_tokens[1:])
            # if there is no text left over
            if not product_description:
                break
            # replacing delinquent text
            product_description = product_description.replace('.', '').replace(',', '').replace('_', '')
            product_description = product_description.replace(':', '').replace(u'\u2014', '').replace(u'\u002A', '')
            product_description = product_description.strip().replace(u'\u201d', '').replace(u'\u2014', '')
            # place our parsed product description in the restructured spreadsheet
            structured_sheet.cell(row=cur_pos+1, column=2).value = product_description
            # if we aren't currently tracking this row as a product description row, add it to the tracked list
            if j not in pd_rows:
                pd_rows.append(j)
            break
        # if the cell doesn't contain a semicolon, it typically doesn't contain a product description
        if ':' not in current_cell.value:
            continue
        # there typically is a product description if it does
        else:
            # remove all parenthesized text
            cell_text = re.sub('\(.*\)', '', current_cell.value)
            # split the cell's text into chunks, separated by comma occurence
            cell_tokens = cell_text.split(',')
            # outsource implementation to method
            product_description = get_description(cell_tokens)
            # if there is no text left over...
            if not product_description:
                continue
            # replacing delinquent text
            product_description = product_description.replace('.', '').replace(',', '').replace('_', '')
            product_description = product_description.replace(':', '').replace(u'\u2014', '').replace(u'\u002A', '')
            product_description = product_description.strip().replace(u'\u201d', '').replace(u'\u2014', '')
            # place our parsed product description in the restructured spreadsheet
            structured_sheet.cell(row=cur_pos+1, column=2).value = product_description
            # if we aren't currently tracking this row as a product description row, add it to the tracked list
            if j not in pd_rows:
                pd_rows.append(j)
            break
    
    # if the cell is empty after the execution of this logic, set the text in the cell equal to "NA"
    if not structured_sheet.cell(row=cur_pos+1, column=2).value:
        structured_sheet.cell(row=cur_pos+1, column=2).value = 'NA'

        
def get_description(tokens):
    ret = ''
    for token in tokens:
        if any(char.isdigit() for char in token) or len(token) == 0 or 'except as noted' in token or 'see' in token:
            continue
        else:
            if ret == '':
                ret = token.strip()
            else:
                ret += ', ' + token.strip()
    return ret.strip().strip(':')


def get_brand_name(sheet, structured_sheet, cur_pos, i):
    # find and store the brand name (if it exists)
    for j in range(i+1, 0, -1):
        # grab data in current cell, in the first column            
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        if j+1 in bn_rows:
            break
        # if the cell contains parenthesized text (not equal to cont.), it must contain brand name information
        if current_cell.value.find('(') != -1 and '(cont.)' not in current_cell.value and 'see also' not in current_cell.value:
            # get the cell's text within parentheses; this is the brand name
            brand_name = current_cell.value[current_cell.value.find("(")+1:current_cell.value.find(")")]
            brand_name = brand_name.replace('.', '').replace('_', '')
            structured_sheet.cell(row=cur_pos+1, column=4).value = brand_name
            if j not in bn_rows:
                bn_rows.append(j)
            break

        elif j in pc_rows or j in pd_rows:
            # there cannot be any brand name information before here
            structured_sheet.cell(row=cur_pos+1, column=4).value = 'NA'
            break

    if not structured_sheet.cell(row=cur_pos+1, column=4).value.strip():
        structured_sheet.cell(row=cur_pos+1, column=4).value = 'NA'


def get_type(sheet, structured_sheet, cur_pos, i):
    # find and store the type (if it exists)
    for j in range(i+1, 0, -1): # check current and all previous rows up to the last product category row
        # grab data in current cell, in the first column
        current_cell = sheet.cell(row=j, column=1)            
        if not current_cell.value:
            continue
        if j in pc_rows or j+1 in bn_rows:
            break
        # if the cell contains parentheses
        if '(' in current_cell.value and j in bn_rows and '(cont.)' not in current_cell.value:
            # get the token with parentheses
            token = current_cell.value.split('(')[0]
            if token != '' and ':' not in token:
                token = token.strip().replace('.', '').replace('_', '')
                type_ = remove_numerics(token)
                structured_sheet.cell(row=cur_pos+1, column=3).value = type_
                break
        # if the cell doesn't contain parentheses and doesn't contain a semicolon
        elif ':' not in current_cell.value and '(' not in current_cell.value:
            type_ = current_cell.value.replace('.', '').replace('_', '').strip()
            type_ = remove_numerics(type_)
            structured_sheet.cell(row=cur_pos+1, column=3).value = type_
            break

    if not structured_sheet.cell(row=cur_pos+1, column=3).value.strip():
        structured_sheet.cell(row=cur_pos+1, column=3).value = 'NA'


def remove_numerics(token):
    ret = ''
    for x in token.split(','):
        if not any(char.isdigit() for char in x):
            ret += x + ','
    ret = ret.strip().strip(',')
    return ret


def get_serving_size(sheet, structured_sheet, cur_pos, i):
    # find and store the serving size (each entry must have one)
    for j in range(i+1, 0, -1): # check all previous rows up to the last product category row
        # grab data in current cell, in the first column            
        current_cell = sheet.cell(row=j, column=1)
        if not current_cell.value:
            continue
        # if the cell contains some numerical data, it will pertain to the serving size
        if any(char.isdigit() for char in current_cell.value):
            # split the cell's text by comma, looking for the token with numerical data
            serving_size = format_serving_size(current_cell.value.split(','))

            if serving_size:
                serving_size = serving_size.strip('.').replace('_', '').replace('*', '').strip()
                structured_sheet.cell(row=cur_pos+1, column=5).value = serving_size
                break
            else:
                continue

        elif j in pd_rows or j in pc_rows:
            # split the cell's text by comma, looking for the token with numerical data
            serving_size = format_serving_size(current_cell.value.split(','))

            if serving_size:
                serving_size = serving_size.strip('.').replace('_', '').replace('*', '').strip()
                structured_sheet.cell(row=cur_pos+1, column=5).value = serving_size
                break
            else:
                continue

                
def format_serving_size(tokens):
    for token in tokens: # loop through each token
        if any(char.isdigit() for char in token) and '(' not in token and u'\N{DEGREE SIGN}' not in token: # looking for the token with the serving size data...
            token = token.strip().strip(':')
            return token.split(':')[0]
        elif any(char.isdigit() for char in token.split(':')[0]) and '(' not in token.split(':')[0] and u'\N{DEGREE SIGN}' not in token:
            return token.split(':')[0].strip()
        
